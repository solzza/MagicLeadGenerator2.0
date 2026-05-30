from __future__ import annotations

import io
import re
import unicodedata
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


RAW_REQUIRED_COLUMNS = [
    "First Name",
    "Last Name",
    "Company Name",
    "Company Domain",
]

RAW_OPTIONAL_COLUMNS = [
    "Email",
    "Email Status",
    "Prospect Position",
    "Prospect Linkedin URL",
]

EXPORT_COLUMNS = [
    "Name",
    "First Name",
    "Last Name",
    "Prospect Position",
    "Prospect Linkedin URL",
    "Company Upsales",
    "Email",
    "Duplicate Match",
    "Issues",
]

REVIEW_COLUMNS = [
    "Name",
    "First Name",
    "Last Name",
    "Prospect Position",
    "Prospect Linkedin URL",
    "Company Upsales",
    "Company Name",
    "Email",
    "Email Status",
    "Company Domain",
    "Match Status",
    "Issues",
    "Duplicate Key",
]


@dataclass(frozen=True)
class CleanOptions:
    prefer_existing_email: bool = False
    fuzzy_threshold: float = 0.92
    remove_title_keywords: tuple[str, ...] = ()


def normalize_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def normalize_key(value: object) -> str:
    text = normalize_text(value).lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9&]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_header(value: object) -> str:
    text = normalize_key(value)
    text = text.replace("linkedin", "linkedin")
    return text


COLUMN_ALIASES = {
    "Name": {
        "name",
        "full name",
        "fullname",
        "contact name",
        "kontaktperson",
        "activity kontaktperson",
        "namn",
        "hela namnet",
    },
    "First Name": {
        "first name",
        "firstname",
        "fornamn",
        "förnamn",
        "given name",
    },
    "Last Name": {
        "last name",
        "lastname",
        "efternamn",
        "surname",
        "family name",
    },
    "Company Name": {
        "company name",
        "company",
        "foretag",
        "företag",
        "organisation",
        "organization",
        "account",
    },
    "Company Domain": {
        "company domain",
        "domain",
        "domän",
        "doman",
        "website",
        "webbplats",
        "company website",
    },
    "Email": {
        "email",
        "e mail",
        "e post",
        "epost",
        "mail",
        "prospect email",
    },
    "Email Status": {
        "email status",
        "email verification",
        "email verification status",
        "verification status",
        "status email",
        "epost status",
        "e post status",
    },
    "Prospect Position": {
        "prospect position",
        "position",
        "title",
        "job title",
        "current title",
        "current job title",
        "current role",
        "professional title",
        "headline",
        "role",
        "job role",
        "work role",
        "befattning",
        "titel",
        "jobbtitel",
        "arbetstitel",
        "arbetsroll",
        "yrkesroll",
    },
    "Prospect Linkedin URL": {
        "prospect linkedin url",
        "linkedin url",
        "linkedin",
        "linkedin profile",
        "profile url",
        "profil url",
        "linkedin länk",
        "linkedin lank",
    },
}


def _series_text(series: pd.Series) -> pd.Series:
    return series.fillna("").astype(str).str.strip()


def _linkedin_score(series: pd.Series, prefer_sales_nav: bool = False) -> float:
    text = _series_text(series).str.lower()
    linkedin = text.str.contains("linkedin.com", regex=False)
    if prefer_sales_nav:
        target = text.str.contains("linkedin.com/sales/", regex=False)
    else:
        target = text.str.contains("linkedin.com/in/", regex=False)
    return float((target.sum() * 2 + linkedin.sum()) / max(len(text), 1))


def _title_score(series: pd.Series) -> float:
    text = _series_text(series)
    if text.empty:
        return 0.0
    lower = text.str.lower()
    numeric_like = text.str.fullmatch(r"\d+([.,]\d+)?", na=False)
    if numeric_like.mean() > 0.5:
        return 0.0
    boolean_like = lower.str.fullmatch(r"true|false|yes|no|ja|nej|0|1", na=False)
    if boolean_like.mean() > 0.5:
        return 0.0
    bad = (
        lower.str.contains("http", regex=False)
        | lower.str.contains("@", regex=False)
        | lower.str.contains("linkedin.com", regex=False)
        | lower.str.fullmatch(r"true|false", case=False, na=False)
    )
    non_empty = text != ""
    likely_title_words = lower.str.contains(
        r"\b(?:manager|owner|lead|director|chef|ansvarig|specialist|consultant|developer|engineer|product|projekt|project|head|vp|ceo|cto|cfo)\b",
        regex=True,
    )
    avg_words = text[non_empty & ~bad].str.split().map(len).mean()
    avg_words = 0 if pd.isna(avg_words) else avg_words
    return float(((non_empty & ~bad).sum() + likely_title_words.sum() * 2 + min(avg_words, 6)) / max(len(text), 1))


def is_probable_title(value: object) -> bool:
    text = normalize_text(value)
    if not text:
        return False
    lower = text.lower()
    if re.fullmatch(r"\d+([.,]\d+)?", lower):
        return False
    if re.fullmatch(r"true|false|yes|no|ja|nej|0|1", lower):
        return False
    if "http" in lower or "linkedin.com" in lower or "@" in lower:
        return False
    if len(text) > 120:
        return False
    return bool(re.search(r"[A-Za-zÅÄÖåäö]", text))


def is_probable_linkedin_profile(value: object) -> bool:
    text = normalize_text(value).lower()
    if not text:
        return False
    return "linkedin.com/in/" in text


def scan_column_quality(raw: pd.DataFrame, mapping: dict[str, str]) -> pd.DataFrame:
    rows = []
    for standard, source in mapping.items():
        series = raw[source]
        non_empty = _series_text(series) != ""
        sample_size = int(non_empty.sum())
        confidence = ""
        warning = ""

        if standard == "Prospect Position":
            score = _title_score(series)
            confidence = f"{min(score, 1):.0%}"
            if score <= 0:
                warning = "Ser inte ut som jobbtitlar"
            elif score < 0.45:
                warning = "Låg säkerhet för jobbtitel"
        elif standard == "Prospect Linkedin URL":
            valid = _series_text(series).map(is_probable_linkedin_profile)
            ratio = float(valid.sum() / max(sample_size, 1))
            confidence = f"{ratio:.0%}"
            if ratio < 0.6:
                warning = "Få LinkedIn-profillänkar hittades"
        elif standard == "Company Domain":
            valid = _series_text(series).map(clean_domain).ne("")
            ratio = float(valid.sum() / max(sample_size, 1))
            confidence = f"{ratio:.0%}"
            if ratio < 0.5:
                warning = "Få giltiga domäner hittades"
        elif standard == "Email":
            valid = _series_text(series).map(is_valid_email)
            ratio = float(valid.sum() / max(sample_size, 1))
            confidence = f"{ratio:.0%}"
        else:
            confidence = f"{float(non_empty.sum() / max(len(series), 1)):.0%}"

        rows.append(
            {
                "Standardkolumn": standard,
                "Hittad kolumn": source,
                "Icke-tomma rader": sample_size,
                "Säkerhet": confidence,
                "Varning": warning,
            }
        )
    return pd.DataFrame(rows)


def infer_column_mapping(raw: pd.DataFrame) -> dict[str, str]:
    mapping: dict[str, str] = {}
    normalized_columns = {column: normalize_header(column) for column in raw.columns}
    used_columns: set[str] = set()

    for target, aliases in COLUMN_ALIASES.items():
        for column, normalized in normalized_columns.items():
            if column in used_columns:
                continue
            if normalized in {normalize_header(alias) for alias in aliases}:
                mapping[target] = column
                used_columns.add(column)
                break

    if "Prospect Linkedin URL" not in mapping:
        candidates = [
            (column, _linkedin_score(raw[column], prefer_sales_nav=False))
            for column in raw.columns
            if column not in used_columns
        ]
        candidates = [candidate for candidate in candidates if candidate[1] > 0]
        if candidates:
            column, _ = max(candidates, key=lambda item: item[1])
            mapping["Prospect Linkedin URL"] = column
            used_columns.add(column)

    if "Prospect Position" not in mapping:
        header_candidates = []
        blocked_title_tokens = {
            "year",
            "years",
            "month",
            "months",
            "duration",
            "tenure",
            "seniority",
            "ar",
            "år",
            "has",
            "new",
            "changed",
            "change",
            "status",
            "is",
            "was",
            "open",
            "profile",
        }
        for column, normalized in normalized_columns.items():
            if column in used_columns:
                continue
            tokens = set(normalized.split())
            if tokens & blocked_title_tokens:
                continue
            score = _title_score(raw[column])
            if any(token in tokens for token in ("position", "title", "titel", "role", "befattning", "jobbtitel", "arbetsroll", "yrkesroll")) and score > 0:
                header_candidates.append((column, score + 2))
        candidates = header_candidates or [
            (column, _title_score(raw[column]))
            for column in raw.columns
            if column not in used_columns
            and not (set(normalized_columns[column].split()) & blocked_title_tokens)
        ]
        candidates = [candidate for candidate in candidates if candidate[1] > 0.35]
        if candidates:
            column, _ = max(candidates, key=lambda item: item[1])
            mapping["Prospect Position"] = column
            used_columns.add(column)

    return mapping


def infer_duplicate_columns(df: pd.DataFrame) -> dict[str, str]:
    mapping = infer_column_mapping(df)
    normalized_columns = {column: normalize_header(column) for column in df.columns}

    if "Name" not in mapping:
        name_aliases = {normalize_header(alias) for alias in COLUMN_ALIASES["Name"]}
        for column, normalized in normalized_columns.items():
            if normalized in name_aliases:
                mapping["Name"] = column
                break

    if "Email" not in mapping:
        for column in df.columns:
            text = _series_text(df[column])
            if text.map(is_valid_email).mean() > 0.4:
                mapping["Email"] = column
                break

    return {key: value for key, value in mapping.items() if key in {"Name", "Email"}}


def mark_duplicates_against(export_df: pd.DataFrame, duplicate_df: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, str]]:
    mapping = infer_duplicate_columns(duplicate_df)
    result = export_df.copy()
    existing_names: set[str] = set()
    existing_emails: set[str] = set()

    if "Name" in mapping:
        existing_names = set(duplicate_df[mapping["Name"]].map(normalize_key))
        existing_names.discard("")
    if "Email" in mapping:
        existing_emails = set(duplicate_df[mapping["Email"]].map(lambda value: normalize_text(value).lower()))
        existing_emails.discard("")

    duplicate_reasons = []
    for _, row in result.iterrows():
        reasons = []
        if normalize_key(row.get("Name", "")) in existing_names:
            reasons.append("Duplicate name")
        if normalize_text(row.get("Email", "")).lower() in existing_emails:
            reasons.append("Duplicate email")
        duplicate_reasons.append(", ".join(reasons))

    result["Duplicate Match"] = duplicate_reasons
    if "Issues" not in result.columns:
        result["Issues"] = ""
    result["Issues"] = [
        ", ".join(part for part in [normalize_text(issue), duplicate] if part)
        for issue, duplicate in zip(result["Issues"], duplicate_reasons)
    ]
    return result, mapping


def has_duplicate_match(value: object) -> bool:
    text = normalize_text(value).lower()
    if not text:
        return False
    duplicate_labels = {"duplicate name", "duplicate email"}
    return any(part.strip() in duplicate_labels for part in text.split(","))


def standardize_raw_columns(raw: pd.DataFrame) -> pd.DataFrame:
    mapping = infer_column_mapping(raw)
    standardized = pd.DataFrame(index=raw.index)
    for target in RAW_REQUIRED_COLUMNS + RAW_OPTIONAL_COLUMNS:
        if target in mapping:
            standardized[target] = raw[mapping[target]]
        elif target in raw.columns:
            standardized[target] = raw[target]
    return standardized


def normalize_email_part(value: object) -> str:
    text = normalize_text(value).lower()
    replacements = {
        "å": "a",
        "ä": "a",
        "ö": "o",
        "é": "e",
        "è": "e",
        "ü": "u",
        " ": "-",
        "'": "",
        "`": "",
        "´": "",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9.-]+", "", text)
    text = re.sub(r"-+", "-", text).strip(".-")
    return text


def clean_domain(value: object) -> str:
    text = normalize_text(value).lower()
    if not text:
        return ""
    text = re.sub(r"^https?://", "", text)
    text = text.split("/")[0]
    text = text.split("?")[0]
    text = text.removeprefix("www.")
    text = text.strip().strip(".")
    return text if "." in text else ""


def is_valid_email(value: object) -> bool:
    text = normalize_text(value).lower()
    return bool(re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", text))


def read_table(uploaded_or_path, sheet_name: str | int | None = 0) -> pd.DataFrame:
    if isinstance(uploaded_or_path, (str, Path)):
        path = Path(uploaded_or_path)
        suffix = path.suffix.lower()
    else:
        suffix = Path(uploaded_or_path.name).suffix.lower()

    if suffix == ".csv":
        return pd.read_csv(uploaded_or_path, sep=None, engine="python", encoding="utf-8-sig")
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        return pd.read_excel(uploaded_or_path, sheet_name=sheet_name)
    raise ValueError(f"Unsupported file type: {suffix}")


def available_sheets(uploaded_or_path) -> list[str]:
    if isinstance(uploaded_or_path, (str, Path)):
        path = Path(uploaded_or_path)
        suffix = path.suffix.lower()
    else:
        path = uploaded_or_path
        suffix = Path(uploaded_or_path.name).suffix.lower()
    if suffix not in {".xlsx", ".xlsm", ".xls"}:
        return []
    excel = pd.ExcelFile(path)
    return excel.sheet_names


def validate_raw_columns(raw: pd.DataFrame) -> list[str]:
    standardized = standardize_raw_columns(raw)
    return [column for column in RAW_REQUIRED_COLUMNS if column not in standardized.columns]


def normalize_raw(raw: pd.DataFrame) -> pd.DataFrame:
    df = standardize_raw_columns(raw)
    for column in RAW_REQUIRED_COLUMNS + RAW_OPTIONAL_COLUMNS:
        if column not in df.columns:
            df[column] = ""
    keep = RAW_REQUIRED_COLUMNS + [c for c in RAW_OPTIONAL_COLUMNS if c in df.columns]
    df = df[keep].copy()
    for column in df.columns:
        df[column] = df[column].map(normalize_text)
    df["Company Domain"] = df["Company Domain"].map(clean_domain)
    return df


def build_company_registry(registry: pd.DataFrame) -> pd.DataFrame:
    df = registry.copy()
    if "Bolag" in df.columns:
        company_col = "Bolag"
    elif "Contact: Företag" in df.columns:
        company_col = "Contact: Företag"
    elif "Company" in df.columns:
        company_col = "Company"
    elif "Företag" in df.columns:
        company_col = "Företag"
    else:
        company_col = df.columns[0]

    if "Uppdaterad domän" in df.columns:
        email_col = "Uppdaterad domän"
    elif "Domain" in df.columns:
        email_col = "Domain"
    elif "Domän" in df.columns:
        email_col = "Domän"
    elif "Doman" in df.columns:
        email_col = "Doman"
    elif "Contact: Email" in df.columns:
        email_col = "Contact: Email"
    elif "Email" in df.columns:
        email_col = "Email"
    else:
        email_col = df.columns[1] if len(df.columns) > 1 else company_col

    out = pd.DataFrame(
        {
            "Company Upsales": df[company_col].map(normalize_text),
            "Email Suffix": df[email_col].map(normalize_text),
        }
    )
    out = out[out["Company Upsales"] != ""].copy()
    out["Company Key"] = out["Company Upsales"].map(normalize_key)
    out["Email Suffix"] = out["Email Suffix"].map(format_email_suffix)
    out["Domain"] = out["Email Suffix"].str.removeprefix("@")
    out = out.drop_duplicates(subset=["Company Key"], keep="first")
    return out


def format_email_suffix(value: object) -> str:
    text = normalize_text(value).lower()
    if not text:
        return ""
    if "@" in text:
        text = text[text.index("@") :]
    else:
        text = "@" + clean_domain(text)
    return text if re.fullmatch(r"@[^@\s]+\.[^@\s]+", text) else ""


def candidate_company_keys(company_name: object) -> Iterable[str]:
    key = normalize_key(company_name)
    if not key:
        return []
    parts = key.split()
    return (" ".join(parts[: len(parts) - cut]) for cut in range(0, len(parts)))


def match_company(company_name: object, registry: pd.DataFrame, fuzzy_threshold: float) -> tuple[str, str, str]:
    if registry.empty:
        return "", "", "no_registry"

    lookup = registry.set_index("Company Key", drop=False)
    for key in candidate_company_keys(company_name):
        if key in lookup.index:
            row = lookup.loc[key]
            if isinstance(row, pd.DataFrame):
                row = row.iloc[0]
            return row["Company Upsales"], row["Email Suffix"], "exact"

    company_key = normalize_key(company_name)
    if not company_key:
        return "", "", "missing_company"

    best_score = 0.0
    best_row = None
    for _, row in registry.iterrows():
        score = SequenceMatcher(None, company_key, row["Company Key"]).ratio()
        if score > best_score:
            best_score = score
            best_row = row

    if best_row is not None and best_score >= fuzzy_threshold:
        return best_row["Company Upsales"], best_row["Email Suffix"], f"fuzzy:{best_score:.2f}"
    return "", "", "not_found"


def match_company_by_domain(domain: object, registry: pd.DataFrame) -> tuple[str, str, str]:
    clean = clean_domain(domain)
    if not clean or registry.empty or "Domain" not in registry.columns:
        return "", "", "not_found"
    hit = registry[registry["Domain"].fillna("").str.lower() == clean.lower()]
    if hit.empty:
        return "", "", "not_found"
    row = hit.iloc[0]
    return row["Company Upsales"], row["Email Suffix"], "domain"


def build_email(first_name: object, last_name: object, domain_or_suffix: object) -> str:
    first = normalize_email_part(first_name)
    last = normalize_email_part(last_name)
    suffix = format_email_suffix(domain_or_suffix)
    if not first or not last or not suffix:
        return ""
    return f"{first}.{last}{suffix}".lower()


def email_status_is_safe(value: object) -> bool:
    status = normalize_key(value)
    safe_terms = {
        "safe",
        "verified",
        "valid",
        "ok",
        "deliverable",
        "confirmed",
        "godkand",
        "godkänd",
    }
    return any(term in status.split() or term == status for term in safe_terms)


def title_is_excluded(title: object, keywords: Iterable[str]) -> bool:
    title_key = normalize_key(title)
    return any(normalize_key(keyword) in title_key for keyword in keywords if normalize_key(keyword))


def clean_leads(raw: pd.DataFrame, registry: pd.DataFrame, options: CleanOptions | None = None) -> pd.DataFrame:
    options = options or CleanOptions()
    raw_df = normalize_raw(raw)
    reg_df = build_company_registry(registry)
    rows = []

    for _, row in raw_df.iterrows():
        issues = []
        company_up, suffix, match_status = match_company(row["Company Name"], reg_df, options.fuzzy_threshold)

        if match_status in {"not_found", "missing_company", "no_registry"}:
            company_by_domain, suffix_by_domain, domain_status = match_company_by_domain(row["Company Domain"], reg_df)
            if domain_status == "domain":
                company_up = company_by_domain
                suffix = suffix_by_domain
                match_status = "domain"
            else:
                company_up = row["Company Name"]
                issues.append("New company")

        domain_suffix = "@" + row["Company Domain"] if row["Company Domain"] else ""
        email = ""
        if is_valid_email(row.get("Email", "")) and (
            email_status_is_safe(row.get("Email Status", "")) or options.prefer_existing_email
        ):
            email = normalize_text(row["Email"]).lower()
        if not email:
            email = build_email(row["First Name"], row["Last Name"], domain_suffix or suffix)

        if not email:
            issues.append("email_missing")
        elif not is_valid_email(email):
            issues.append("email_invalid")

        if not row["Company Domain"] and not suffix:
            issues.append("domain_missing")

        if row.get("Prospect Position", "") and not is_probable_title(row.get("Prospect Position", "")):
            issues.append("title_invalid")
        if not row.get("Prospect Position", ""):
            issues.append("title_missing")
        if row.get("Prospect Linkedin URL", "") and not is_probable_linkedin_profile(row.get("Prospect Linkedin URL", "")):
            issues.append("linkedin_invalid")
        if not row.get("Prospect Linkedin URL", ""):
            issues.append("linkedin_missing")

        if title_is_excluded(row.get("Prospect Position", ""), options.remove_title_keywords):
            issues.append("title_excluded")

        full_name = f"{normalize_text(row['First Name'])} {normalize_text(row['Last Name'])}".strip()

        out = {
            "Name": full_name,
            "First Name": row["First Name"],
            "Last Name": row["Last Name"],
            "Prospect Position": row.get("Prospect Position", ""),
            "Prospect Linkedin URL": row.get("Prospect Linkedin URL", ""),
            "Company Upsales": company_up,
            "Company Name": row["Company Name"],
            "Email": email,
            "Email Status": row.get("Email Status", ""),
            "Company Domain": row["Company Domain"],
            "Match Status": match_status,
            "Issues": ", ".join(issues),
        }
        rows.append(out)

    cleaned = pd.DataFrame(rows)
    duplicate_key = cleaned["Email"].where(cleaned["Email"] != "", cleaned["Prospect Linkedin URL"])
    cleaned["Duplicate Key"] = duplicate_key.where(duplicate_key.duplicated(keep=False), "")
    return cleaned[REVIEW_COLUMNS]


def to_import_columns(df: pd.DataFrame) -> pd.DataFrame:
    export = df.copy()
    first = export.get("First Name", pd.Series("", index=export.index)).fillna("").astype(str).str.strip()
    last = export.get("Last Name", pd.Series("", index=export.index)).fillna("").astype(str).str.strip()
    built_name = (first + " " + last).str.strip()
    if "Name" not in export.columns:
        export["Name"] = built_name
    else:
        export["Name"] = export["Name"].fillna("").astype(str).str.strip()
        export.loc[export["Name"] == "", "Name"] = built_name[export["Name"] == ""]
    for column in EXPORT_COLUMNS:
        if column not in export.columns:
            export[column] = ""
    return export[EXPORT_COLUMNS]


def dataframe_to_xlsx_bytes(df: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Upsales Import")
        ws = writer.book["Upsales Import"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        header_fill = PatternFill("solid", fgColor="1F2428")
        header_font = Font(bold=True, size=16, color="D8BE6A")
        body_font = Font(size=12, color="111111")
        duplicate_fill = PatternFill("solid", fgColor="FFF8E1")
        duplicate_side = Side(style="medium", color="C08A00")
        medium_black = Side(style="medium", color="000000")

        ws.row_dimensions[1].height = 36
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = Border(
                left=medium_black,
                right=medium_black,
                top=medium_black,
                bottom=medium_black,
            )

        duplicate_col = None
        for cell in ws[1]:
            if cell.value == "Duplicate Match":
                duplicate_col = cell.column

        duplicate_rows = {}
        for row_number in range(2, ws.max_row + 1):
            duplicate_value = ws.cell(row=row_number, column=duplicate_col).value if duplicate_col else ""
            duplicate_rows[row_number] = has_duplicate_match(duplicate_value)

        for row in ws.iter_rows(min_row=2):
            duplicate_value = ws.cell(row=row[0].row, column=duplicate_col).value if duplicate_col else ""
            row_number = row[0].row
            is_duplicate = duplicate_rows[row_number]
            previous_duplicate = duplicate_rows.get(row_number - 1, False)
            next_duplicate = duplicate_rows.get(row_number + 1, False)
            for cell in row:
                if is_duplicate:
                    cell.fill = duplicate_fill
                cell.font = body_font
                cell.alignment = Alignment(horizontal="left", vertical="top")
                if is_duplicate:
                    cell.border = Border(
                        top=duplicate_side if not previous_duplicate else None,
                        bottom=duplicate_side if not next_duplicate else None,
                    )
                else:
                    cell.border = Border()

        for column_cells in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 14), 52)
    return output.getvalue()
